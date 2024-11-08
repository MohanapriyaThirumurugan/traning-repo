# Day 1:
# CSV to Excel Conversion: Explored methods to convert CSV files into Excel format using libraries like pandas and openpyxl. The focus was on handling different data types and ensuring that the structure of the data remained intact after conversion.
# Excel to CSV Conversion: Learned how to convert Excel files back into CSV format. This was essential for ensuring that large Excel files could be used for data processing in formats that are lightweight and compatible with various data processing tools.
# Excel to Text Conversion: Practiced converting Excel data into plain text format. This exercise helped in understanding how to deal with unstructured data and how it can be saved or utilized in different text-based applications.


import csv
rows = []

# read file 
with open("TVAQueries.txt", "r", encoding="utf-8") as file:
    
    content = file.readlines() # read all line 
    temp = []  
    for i in content: 
        temp.append(i)
        if len(temp) == 100:
            rows.append(temp)  
            temp = []   

print(len(rows))       

# Open a new CSV file to write the data
with open("fine4.csv", "w", encoding="utf-8", newline='') as file1:
    
    csv_writer = csv.writer(file1)

    csv_writer.writerows(rows) # write row is function

print("added ")

# fetch data from  xl to text


import openpyxl

# Step 1: Load the Excel file
excel_file_path = 'filecsv.xlsx'
workbook = openpyxl.load_workbook(excel_file_path)

# Step 2: Select the active sheet
sheet = workbook.active

# Step 3: Open the text file in write mode
output_file_path = 'output_file.txt'
with open(output_file_path, 'w') as file:
    
    # Step 4: Loop through rows and columns manually
    for i in range(1, sheet.max_row + 1):  # Loop through each row
        row_data = []  # Store cell values of the current row
        for j in range(1, sheet.max_column + 1):  # Loop through each column
            cell_value = sheet.cell(row=i, column=j).value  # Get the value of the cell
            print(cell_value)
            row_data.append(str(cell_value))  # Add the value to the row data
        
        # Step 5: Write the row data to the text file
        file.write('\t'.join(row_data) + '\n')  # Write the row to the text file, separated by tabs


        # acces full coloum
        import openpyxl

# Load the Excel file
excel_file_path = 'filecsv.xlsx'
workbook = openpyxl.load_workbook(excel_file_path)
sheet = workbook.active  # Select the active sheet

# Specify the column you want to access
specific_column = 3  # Example: Access the third column (C)

# List to hold values of the specified column
column_data = []  

# Loop through each row in the specified column
for i in range(1, sheet.max_row + 1):  # Loop from the first row to the last row
    cell_value = sheet.cell(row=i, column=specific_column).value  # Get the value of the cell
    column_data.append(cell_value)  # Append the value to the column_data list

# Print the values in the specified column
print(f'Values in Column {specific_column}: {column_data}')

# access full row
import openpyxl

# Load the Excel file
excel_file_path = 'filecsv.xlsx'
workbook = openpyxl.load_workbook(excel_file_path)
sheet = workbook.active  # Select the active sheet

# Specify the column you want to access
specific_column = 3  # Example: Access the third column (C)

# List to hold values of the specified column
column_data = []  

# Loop through each row in the specified column
for i in range(1, sheet.max_row + 1):  # Loop from the first row to the last row
    cell_value = sheet.cell(row=i, column=specific_column).value  # Get the value of the cell
    column_data.append(cell_value)  # Append the value to the column_data list

# Print the values in the specified column
print(f'Values in Column {specific_column}: {column_data}')



# like how acces csv to xl formate

 

import csv  # Import the csv module
import openpyxl  # Import the openpyxl library

# Step 1: Load the CSV file
csv_file_path = 'input_file.csv'  # Specify the path to your CSV file
excel_file_path = 'output_file.xlsx'  # Specify the path for the output Excel file

# Step 2: Create a new Excel workbook and select the active sheet
workbook = openpyxl.Workbook()  # Create a new workbook
sheet = workbook.active  # Get the active sheet

# Step 3: Read the CSV file and write to the Excel sheet
with open(csv_file_path, newline='', encoding='utf-8') as csvfile:
    csv_reader = csv.reader(csvfile)  # Create a CSV reader object
    for row in csv_reader:  # Loop through each row in the CSV
        sheet.append(row)  # Append the row to the Excel sheet

# Step 4: Save the Excel workbook
workbook.save(excel_file_path)  # Save the workbook to the specified path

print(f'Converted {csv_file_path} to {excel_file_path}')

